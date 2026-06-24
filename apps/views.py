import io
from django.db.models import Sum, F, DecimalField, ExpressionWrapper
from django.http import HttpResponse
from drf_spectacular.utils import extend_schema, extend_schema_view
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.viewsets import ModelViewSet

from apps.models import Product, Stock, Sale, Category
from apps.permissions import IsOperatorOrReadOnly, IsManagement
from apps.serializers import (ProductSerializer, ProductOperatorSerializer,
                               StockSerializer, SaleSerializer,
                               LoginSerializer, RegisterOperatorSerializer,
                               CategorySerializer)


@extend_schema(
    summary="Login — JWT token olish",
    description=(
        "Username va parol bilan kirish. Muvaffaqiyatli bo'lsa "
        "`access` va `refresh` tokenlar qaytariladi.\n\n"
        "`Authorization: Bearer <access_token>` sarlavhasida ishlating."
    ),
    tags=["Auth"],
    request=LoginSerializer,
    responses={200: LoginSerializer},
    auth=[],
)
@api_view(['POST'])
@permission_classes([AllowAny])
def login(request):
    serializer = LoginSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    return Response(serializer.data)


@extend_schema(
    summary="Yangi operator yaratish (faqat Management)",
    description="Faqat `MANAGEMENT` roli yangi `OPERATOR` foydalanuvchi qo'sha oladi.",
    tags=["Auth"],
    request=RegisterOperatorSerializer,
    responses={201: RegisterOperatorSerializer},
)
@api_view(['POST'])
@permission_classes([IsManagement])
def register_operator(request):
    serializer = RegisterOperatorSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response(serializer.data, status=status.HTTP_201_CREATED)


@extend_schema_view(
    list=extend_schema(
        summary="Kategoriyalar daraxti",
        description="Faqat root (ota yo'q) kategoriyalar, ichida children rekursiv.",
        tags=["Categories"],
    ),
    retrieve=extend_schema(summary="Kategoriya ma'lumoti", tags=["Categories"]),
    create=extend_schema(summary="Yangi kategoriya (Operator)", tags=["Categories"]),
    update=extend_schema(summary="Kategoriyani yangilash (Operator)", tags=["Categories"]),
    partial_update=extend_schema(summary="Kategoriyani qisman yangilash (Operator)", tags=["Categories"]),
    destroy=extend_schema(summary="Kategoriyani o'chirish (Operator)", tags=["Categories"]),
)
class CategoryViewSet(ModelViewSet):
    serializer_class = CategorySerializer
    permission_classes = (IsOperatorOrReadOnly,)
    search_fields = ('name',)

    def get_queryset(self):
        if self.action == 'list':
            return Category.objects.root_nodes().prefetch_related(
                'children__children__children'
            )
        return Category.objects.all()


@extend_schema_view(
    list=extend_schema(
        summary="Barcha mahsulotlar ro'yxati",
        description="Qidiruv: `?search=nomi/modeli/seriya`. Filtr: `?ordering=purchase_price`.",
        tags=["Products"],
    ),
    retrieve=extend_schema(summary="Mahsulot ma'lumoti", tags=["Products"]),
    create=extend_schema(summary="Yangi mahsulot qo'shish (Operator)", tags=["Products"]),
    update=extend_schema(summary="Mahsulotni to'liq yangilash (Operator)", tags=["Products"]),
    partial_update=extend_schema(summary="Mahsulotni qisman yangilash (Operator)", tags=["Products"]),
    destroy=extend_schema(summary="Mahsulotni o'chirish (Operator)", tags=["Products"]),
)
class ProductViewSet(ModelViewSet):
    queryset = Product.objects.all()
    permission_classes = (IsOperatorOrReadOnly,)
    search_fields = ('name', 'model', 'serial_number')
    ordering_fields = ('name', 'purchase_price', 'created_at')

    def get_serializer_class(self):
        user = self.request.user
        if user.is_authenticated and user.is_management:
            return ProductSerializer
        return ProductOperatorSerializer


@extend_schema_view(
    list=extend_schema(
        summary="Ombordagi qoldiqlar",
        description="Filtr: `?product=1`, `?warehouse_location=A1`.",
        tags=["Stock"],
    ),
    retrieve=extend_schema(summary="Qoldiq ma'lumoti", tags=["Stock"]),
    create=extend_schema(summary="Yangi qoldiq yozuvi (Operator)", tags=["Stock"]),
    update=extend_schema(summary="Qoldiqni to'liq yangilash (Operator)", tags=["Stock"]),
    partial_update=extend_schema(summary="Qoldiqni qisman yangilash (Operator)", tags=["Stock"]),
    destroy=extend_schema(summary="Qoldiq yozuvini o'chirish (Operator)", tags=["Stock"]),
)
class StockViewSet(ModelViewSet):
    queryset = Stock.objects.select_related('product')
    serializer_class = StockSerializer
    permission_classes = (IsOperatorOrReadOnly,)
    filterset_fields = ('product', 'warehouse_location')
    search_fields = ('product__name', 'product__serial_number', 'warehouse_location')


@extend_schema_view(
    list=extend_schema(
        summary="Sotuvlar ro'yxati",
        description="Filtr: `?product=1`, `?sold_date=2024-01-15`.",
        tags=["Sales"],
    ),
    retrieve=extend_schema(summary="Sotuv ma'lumoti", tags=["Sales"]),
    create=extend_schema(
        summary="Yangi sotuv (Operator)",
        description=(
            "Sotuv yaratilganda ombor qoldig'i avtomatik kamayadi (FIFO). "
            "Agar so'ralgan miqdor qoldiqdan ko'p bo'lsa — xato qaytariladi."
        ),
        tags=["Sales"],
    ),
    update=extend_schema(summary="Sotuvni to'liq yangilash (Operator)", tags=["Sales"]),
    partial_update=extend_schema(summary="Sotuvni qisman yangilash (Operator)", tags=["Sales"]),
    destroy=extend_schema(summary="Sotuv yozuvini o'chirish (Operator)", tags=["Sales"]),
)
class SaleViewSet(ModelViewSet):
    queryset = Sale.objects.select_related('product')
    serializer_class = SaleSerializer
    permission_classes = (IsOperatorOrReadOnly,)
    filterset_fields = ('product', 'sold_date')
    search_fields = ('product__name', 'sold_to', 'destination')
    ordering_fields = ('sold_date', 'sold_price', 'quantity')



@extend_schema(
    summary="Bosh hisobot (Management)",
    description=(
        "**Faqat Management roli uchun.**\n\n"
        "Umumiy daromad, foyda, sotilgan birliklar soni va "
        "mahsulot bo'yicha kesimni qaytaradi."
    ),
    tags=["Reports"],
    responses={
        200: {
            "type": "object",
            "properties": {
                "sales_count":       {"type": "integer"},
                "total_revenue":     {"type": "number"},
                "total_profit":      {"type": "number"},
                "total_units_sold":  {"type": "integer"},
                "total_stock_units": {"type": "integer"},
                "products_count":    {"type": "integer"},
                "by_product":        {"type": "array", "items": {"type": "object"}},
            },
        }
    },
)
@api_view(['GET'])
@permission_classes([IsManagement])
def reports(request):
    profit_expr = ExpressionWrapper(
        (F('sold_price') - F('product__purchase_price')) * F('quantity'),
        output_field=DecimalField(max_digits=18, decimal_places=2),
    )
    revenue_expr = ExpressionWrapper(
        F('sold_price') * F('quantity'),
        output_field=DecimalField(max_digits=18, decimal_places=2),
    )

    sales = Sale.objects.all()
    totals = sales.aggregate(
        total_revenue=Sum(revenue_expr),
        total_profit=Sum(profit_expr),
        total_units_sold=Sum('quantity'),
    )

    by_product = list(
        sales.values('product', 'product__name')
        .annotate(
            revenue=Sum(revenue_expr),
            profit=Sum(profit_expr),
            units_sold=Sum('quantity'),
        )
        .order_by('-profit')
    )

    return Response({
        'sales_count':       sales.count(),
        'total_revenue':     totals['total_revenue'] or 0,
        'total_profit':      totals['total_profit'] or 0,
        'total_units_sold':  totals['total_units_sold'] or 0,
        'total_stock_units': Stock.objects.aggregate(t=Sum('quantity'))['t'] or 0,
        'products_count':    Product.objects.count(),
        'by_product':        by_product,
    })


# ─────────────────────────────────────────────
#  EXCEL EXPORT YORDAMCHI
# ─────────────────────────────────────────────
def _make_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def _header_row(ws, cols):
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1F4E79')
    for col_num, title in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_num, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')


# ─────────────────────────────────────────────
#  SOTUVLAR EXCEL
# ─────────────────────────────────────────────
@extend_schema(
    summary="Sotuvlarni Excel yuklab olish",
    tags=["Export"],
    responses={200: None},
)
@api_view(['GET'])
@permission_classes([IsManagement])
def export_sales(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sotuvlar'

    cols = ['#', 'Mahsulot', 'Miqdor', 'Sotuv narxi', 'Jami summa',
            'Foyda', 'Xaridor', 'Qayerga ketdi', 'Sana', 'Izoh']
    _header_row(ws, cols)

    profit_expr = ExpressionWrapper(
        (F('sold_price') - F('product__purchase_price')) * F('quantity'),
        output_field=DecimalField(max_digits=18, decimal_places=2),
    )
    sales = (Sale.objects
             .select_related('product')
             .annotate(profit_val=profit_expr)
             .order_by('-sold_date'))

    for i, s in enumerate(sales, start=1):
        ws.append([
            i,
            str(s.product),
            s.quantity,
            float(s.sold_price),
            float(s.sold_price * s.quantity),
            float(s.profit_val or 0),
            s.sold_to or '',
            s.destination or '',
            s.sold_date.strftime('%Y-%m-%d'),
            s.comment or '',
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            len(str(col[0].value or '')), 12
        )

    return _make_response(wb, 'sotuvlar.xlsx')


# ─────────────────────────────────────────────
#  OMBOR QOLDIQLARI EXCEL
# ─────────────────────────────────────────────
@extend_schema(
    summary="Ombor qoldiqlarini Excel yuklab olish",
    tags=["Export"],
    responses={200: None},
)
@api_view(['GET'])
@permission_classes([IsManagement])
def export_stock(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Ombor qoldiqlari'

    cols = ['#', 'Mahsulot', 'Kategoriya', 'Model', 'Seriya raqami',
            'Qayerdan keldi', 'Lokatsiya', 'Miqdor', 'Olish narxi']
    _header_row(ws, cols)

    stocks = (Stock.objects
              .select_related('product', 'product__category')
              .order_by('product__name', 'warehouse_location'))

    for i, st in enumerate(stocks, start=1):
        p = st.product
        ws.append([
            i,
            p.name,
            str(p.category) if p.category else '',
            p.model or '',
            p.serial_number,
            p.source or '',
            st.warehouse_location,
            st.quantity,
            float(p.purchase_price),
        ])

        # Miqdor 0 bo'lsa — qizil rang
        if st.quantity == 0:
            from openpyxl.styles import PatternFill as PF
            for cell in ws[ws.max_row]:
                cell.fill = PF('solid', fgColor='FFCCCC')

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            len(str(col[0].value or '')), 12
        )

    return _make_response(wb, 'ombor_qoldiqlari.xlsx')
