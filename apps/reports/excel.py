"""
Excel export helpers. Each function returns an HttpResponse with an .xlsx file.
"""
import io
from datetime import date, datetime
from decimal import Decimal

import openpyxl
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill('solid', fgColor='2E75B6')
TITLE_FONT = Font(bold=True, size=14, color='1E1B2E')
META_FONT = Font(size=10, color='5B6472')
CENTER = Alignment(horizontal='center', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin', color='E5E7EB'),
    right=Side(style='thin', color='E5E7EB'),
    top=Side(style='thin', color='E5E7EB'),
    bottom=Side(style='thin', color='E5E7EB'),
)


def _period_label(date_from=None, date_to=None) -> str:
    if not date_from and not date_to:
        return 'Barcha davr'
    if date_from == date_to:
        return str(date_from)
    return f'{date_from or "…"} — {date_to or "…"}'


def _create_workbook(title: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]
    return wb, ws


def _autosize_columns(ws, min_width=10, max_width=42):
    for col_idx, column_cells in enumerate(ws.columns, 1):
        length = min_width
        for cell in column_cells:
            if cell.value is None:
                continue
            length = max(length, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[get_column_letter(col_idx)].width = length


def _write_report_meta(ws, report_title: str, date_from=None, date_to=None,
                       extra: list[tuple[str, str]] | None = None):
    ws.append([report_title])
    ws[ws.max_row][0].font = TITLE_FONT
    ws.append(['Davr', _period_label(date_from, date_to)])
    ws.append(['Yaratilgan', timezone.localtime().strftime('%Y-%m-%d %H:%M')])
    for key, value in extra or []:
        ws.append([key, value])
    for row in range(2, ws.max_row + 1):
        ws[row][0].font = META_FONT
    ws.append([])


def _write_header(ws, columns: list[str], start_row: int | None = None):
    if start_row:
        ws.insert_rows(start_row)
        row_idx = start_row
    else:
        ws.append(columns)
        row_idx = ws.max_row
        for col_idx, label in enumerate(columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=label)
    for col_idx, label in enumerate(columns, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    return row_idx + 1


def _response(wb, filename: str) -> HttpResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _filename(prefix: str, date_from=None, date_to=None) -> str:
    period = _period_label(date_from, date_to).replace(' — ', '_').replace(' ', '')
    if period == 'Barchadavr':
        period = date.today().isoformat()
    return f'{prefix}_{period}.xlsx'


def export_sales(queryset, date_from=None, date_to=None) -> HttpResponse:
    wb, ws = _create_workbook('Sotuvlar')
    total_sum = Decimal('0')
    qty_sum = 0
    _write_report_meta(ws, 'Sotuvlar hisoboti', date_from, date_to)
    header_row = ws.max_row + 1
    columns = [
        '№', 'Mahsulot', 'Kategoriya', 'Miqdor', 'Sotuv narxi',
        'Jami summa', 'Qayerga ketdi', 'Mijoz', 'Sana', 'Izoh',
    ]
    ws.append(columns)
    for col_idx, label in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    for i, sale in enumerate(queryset, 1):
        line_total = sale.sold_price * sale.quantity
        total_sum += line_total
        qty_sum += sale.quantity or 0
        ws.append([
            i,
            str(sale.product),
            str(sale.product.category) if sale.product.category else '',
            sale.quantity,
            float(sale.sold_price),
            float(line_total),
            sale.destination or '',
            sale.sold_to or '',
            sale.sold_date.isoformat() if sale.sold_date else '',
            sale.comment or '',
        ])

    ws.append([])
    ws.append(['', '', 'JAMI', qty_sum, '', float(total_sum)])
    ws[ws.max_row][2].font = Font(bold=True)
    ws[ws.max_row][5].font = Font(bold=True)
    _autosize_columns(ws)
    return _response(wb, _filename('sotuvlar', date_from, date_to))


def export_stock(queryset) -> HttpResponse:
    from apps.warehouse.models import VatPercent

    wb, ws = _create_workbook('Ombor holati')
    _write_report_meta(ws, 'Ombor holati hisoboti',
                       extra=[('Holat', 'Joriy qoldiqlar (snapshot)')])
    columns = [
        '№', 'Mahsulot nomi', 'Kategoriya', 'Seriya raqami', 'Shtrix kod',
        'O\'lchov birligi', 'Qoldiq', 'Bron', 'Mavjud', 'Omborxona',
        'Kelish narxi', 'Sotuv narxi', 'Yetkazish narxi', 'QQS %',
        'QQS miqdori (qoldiq)', 'Jami (qoldiq×sotuv)', 'Minimal qoldiq',
        'Holat', 'Manba',
    ]
    header_row = ws.max_row + 1
    ws.append(columns)
    for col_idx, label in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    for i, stock in enumerate(queryset, 1):
        product = stock.product
        qty = stock.quantity or 0
        reserved = stock.reserved_quantity or 0
        available = qty - reserved
        sell = product.selling_price
        delivery = product.delivery_price
        vat_rate = VatPercent.rate(product.vat_percent)
        vat_label = dict(VatPercent.choices).get(product.vat_percent, product.vat_percent or '')
        line_base = (Decimal(sell or 0) * Decimal(qty)) if sell is not None else Decimal('0')
        vat_amount = (line_base * vat_rate / Decimal('100')).quantize(Decimal('0.01')) if sell else ''
        total = (line_base + vat_amount).quantize(Decimal('0.01')) if sell else ''
        ws.append([
            i,
            product.name,
            str(product.category) if product.category else '',
            product.serial_number,
            product.barcode or '',
            product.get_unit_display(),
            qty,
            reserved,
            available,
            stock.warehouse_location,
            float(product.purchase_price) if product.purchase_price is not None else '',
            float(sell) if sell is not None else '',
            float(delivery) if delivery is not None else '',
            vat_label,
            float(vat_amount) if vat_amount != '' else '',
            float(total) if total != '' else '',
            product.min_quantity,
            product.stock_status,
            product.source or '',
        ])
    ws.append([])
    ws.append(['', f'Jami pozitsiyalar: {queryset.count()}'])
    _autosize_columns(ws)
    return _response(wb, _filename('ombor'))


def export_expenses(queryset, date_from=None, date_to=None) -> HttpResponse:
    wb, ws = _create_workbook('Rasxodlar')
    total_uzs = sum(
        (e.amount for e in queryset if e.currency == 'UZS'), Decimal('0'))
    total_usd = sum(
        (e.amount for e in queryset if e.currency == 'USD'), Decimal('0'))
    _write_report_meta(ws, 'Xarajatlar hisoboti', date_from, date_to,
                       extra=[
                           ('Jami UZS', f'{float(total_uzs):,.0f}'),
                           ('Jami USD', f'{float(total_usd):,.2f}'),
                       ])
    columns = [
        '№', 'Toifa', 'Tur', 'Summa', 'Valyuta',
        'Sana', 'Mas\'ul', 'Izoh', 'Import ID',
    ]
    header_row = ws.max_row + 1
    ws.append(columns)
    for col_idx, label in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    for i, exp in enumerate(queryset, 1):
        ws.append([
            i,
            str(exp.expense_type),
            str(exp.sub_type) if exp.sub_type else '',
            float(exp.amount),
            exp.currency,
            exp.date.isoformat(),
            str(exp.responsible) if exp.responsible else '',
            exp.comment or '',
            exp.zakaz_id or '',
        ])
    _autosize_columns(ws)
    return _response(wb, _filename('xarajatlar', date_from, date_to))


def export_kassa_ledger(date_from=None, date_to=None) -> HttpResponse:
    from apps.cash.ledger import build_ledger_entries

    entries = build_ledger_entries(date_from=date_from, date_to=date_to)

    wb, ws = _create_workbook('Kassa')
    source_labels = {'sale': 'Sotuv', 'order': 'Buyurtma', 'import': 'Import'}
    in_sum = Decimal('0')
    out_sum = Decimal('0')
    in_uzs = Decimal('0')
    out_uzs = Decimal('0')
    for row in entries:
        amount = Decimal(str(row['amount']))
        cur = row.get('currency') or 'UZS'
        if row['kind'] == 'out':
            out_sum += amount
            if cur == 'UZS':
                out_uzs += amount
        else:
            in_sum += amount
            if cur == 'UZS':
                in_uzs += amount

    _write_report_meta(ws, 'Kassa harakatlari', date_from, date_to, extra=[
        ('Tushum UZS', f'{float(in_uzs):,.0f}'),
        ('Import chiqim UZS', f'{float(out_uzs):,.0f}'),
        ('Kassa balansi UZS', f'{float(in_uzs - out_uzs):,.0f}'),
    ])
    columns = [
        '№', 'Turi', 'Manba', 'Izoh', 'Kim / Etkazuvchi',
        'Summa', 'Valyuta', 'Sana',
    ]
    header_row = ws.max_row + 1
    ws.append(columns)
    for col_idx, label in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    for i, row in enumerate(entries, 1):
        amount = Decimal(str(row['amount']))
        signed = -float(amount) if row['kind'] == 'out' else float(amount)
        ws.append([
            i,
            'Chiqim' if row['kind'] == 'out' else 'Tushum',
            source_labels.get(row['source'], row['source']),
            row.get('label') or '',
            row.get('client_name') or '',
            signed,
            row.get('currency') or 'UZS',
            row.get('date') or '',
        ])

    ws.append([])
    ws.append(['', '', '', 'Jami tushum', '', float(in_sum)])
    ws.append(['', '', '', 'Jami chiqim', '', float(-out_sum)])
    ws.append(['', '', '', 'Balans', '', float(in_sum - out_sum)])
    _autosize_columns(ws)
    return _response(wb, _filename('kassa', date_from, date_to))


def export_imports(queryset, date_from=None, date_to=None) -> HttpResponse:
    wb, ws = _create_workbook('Import')
    total_uzs = Decimal('0')
    total_usd = Decimal('0')
    for z in queryset:
        if z.total is None:
            continue
        if z.currency == 'USD':
            total_usd += z.total
        else:
            total_uzs += z.total

    _write_report_meta(ws, 'Import (zakaz) hisoboti', date_from, date_to, extra=[
        ('Jami UZS', f'{float(total_uzs):,.0f}'),
        ('Jami USD', f'{float(total_usd):,.2f}'),
    ])
    columns = [
        '№', 'ID', 'Mahsulot', 'Miqdor', 'Birlik narxi', 'Jami',
        'Valyuta', 'To\'lov holati', 'To\'langan', 'Etkazuvchi',
        'Shartnoma', 'Holati', 'Yaratilgan',
    ]
    header_row = ws.max_row + 1
    ws.append(columns)
    for col_idx, label in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    pay_labels = {'unpaid': 'To\'lanmagan', 'prepaid': 'Oldindan to\'lov', 'paid': 'To\'langan'}
    for i, z in enumerate(queryset, 1):
        ws.append([
            i,
            z.pk,
            str(z.product),
            z.quantity,
            float(z.unit_price) if z.unit_price is not None else '',
            float(z.total) if z.total is not None else '',
            z.currency,
            pay_labels.get(z.payment_status, z.payment_status),
            float(z.paid_amount or 0),
            z.supplier or '',
            z.contract_number or '',
            z.get_status_display(),
            z.created_at.strftime('%Y-%m-%d') if z.created_at else '',
        ])
    _autosize_columns(ws)
    return _response(wb, _filename('import', date_from, date_to))


# Orqaga moslik — eski payments export kassa jurnaliga yo'naltiriladi
def export_payments(queryset=None, date_from=None, date_to=None) -> HttpResponse:
    return export_kassa_ledger(date_from=date_from, date_to=date_to)
